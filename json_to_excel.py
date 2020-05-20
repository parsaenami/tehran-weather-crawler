import pandas as pd
import datetime
import json
import os

def json_to_excel(json_data: dict):

    dates = []
    lowest_temp_data = []
    highest_temp_data = []
    avg_temp_data = []
    weather_status_data = []
    wind_power_data = []
    precipitation_data = []

    for y in json_data.keys():
        for m in json_data[y].keys():
            for d in json_data[y][m].keys():
                dates.append(str(datetime.date(int(y), int(m), int(d))))
                lowest_temp_data.append(json_data[y][m][d]["lowest_temp"])
                highest_temp_data.append(json_data[y][m][d]["highest_temp"])
                avg_temp_data.append(json_data[y][m][d]["avg_temp"])
                weather_status_data.append(json_data[y][m][d]["weather_status"])
                wind_power_data.append(json_data[y][m][d]["wind_power"])
                precipitation_data.append(json_data[y][m][d]["precipitation"])

    data = {
            "dates": dates,
            "lowest_temp_data": lowest_temp_data,
            "highest_temp_data": highest_temp_data,
            "avg_temp_data": avg_temp_data,
            "weather_status_data": weather_status_data,
            "wind_power_data": wind_power_data,
            "precipitation_data": precipitation_data,
        }
    cols = [
            "dates",
            "lowest_temp_data",
            "highest_temp_data",
            "avg_temp_data",
            "weather_status_data",
            "wind_power_data",
            "precipitation_data",
        ]
        
    df = pd.DataFrame(data, columns=cols)
    # df.to_excel(r"test.xlsx", index=False, header=True)
    
    if os.path.exists('weather_log.xlsx'):
        append_df_to_excel('weather_log.xlsx', df, header=None, index=False)
    else:
        append_df_to_excel('weather_log.xlsx', df, header=True, index=False)


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


if __name__ == "__main__":

    with open('weather_log.json') as f:
        info = json.load(f)

    json_to_excel(info)

    print('done')
