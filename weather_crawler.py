import json
import requests
import time
import datetime
import schedule
import os
import sys
import pandas as pd

from bs4 import BeautifulSoup
from collections import Counter
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException

os.system('color')
yb = '\x1b[0;30;43m'
by = '\x1b[2;33;40m'
rb = '\x1b[0;30;41m'
gb = '\x1b[0;30;42m'
pb = '\x1b[5;30;45m'
end = '\x1b[0m'

dates = []
lowest_temp_data = []
highest_temp_data = []
avg_temp_data = []
weather_status_data = []
wind_power_data = []
precipitation_data = []

result = {}
months = {
    1: "january",
    2: "february",
    3: "march",
    4: "april",
    5: "may",
    6: "june",
    7: "july",
    8: "august",
    9: "september",
    10: "october",
    11: "november",
    12: "december",
}

CHROMEDRIVER_PATH = './drivers/chromedriver'
options = Options()
options.headless = True
options.add_argument('--no-sandbox')

user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36"
options.add_argument('user-agent={0}'.format(user_agent))

experimentalFlags = ['same-site-by-default-cookies@1', 'cookies-without-same-site-must-be-secure@1']
chromeLocalStatePrefs = {'browser.enabled_labs_experiments': experimentalFlags}
options.add_experimental_option('localState', chromeLocalStatePrefs)

driver = webdriver.Chrome(CHROMEDRIVER_PATH, options=options)
driver.minimize_window()
driver.execute_script("return navigator.userAgent")

delay = 3  # seconds


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def excel_writer():
    data = {
        "dates": dates,
        "lowest_temp_data": lowest_temp_data,
        "highest_temp_data": highest_temp_data,
        "avg_temp_data": avg_temp_data,
        "weather_status_data": weather_status_data,
        "wind_power_data": wind_power_data,
        "precipitation_data": precipitation_data,
    }
    cols = [
        "dates",
        "lowest_temp_data",
        "highest_temp_data",
        "avg_temp_data",
        "weather_status_data",
        "wind_power_data",
        "precipitation_data",
    ]
    df = pd.DataFrame(data, columns=cols)
    # df.to_excel(r"weather_log.xlsx", index=True, header=True)

    if os.path.exists('weather_log.xlsx'):
        append_df_to_excel('weather_log.xlsx', df, header=None, index=False)
    else:
        append_df_to_excel('weather_log.xlsx', df, header=True, index=False)



def is_int(maybe_int) -> bool:
    try:
        int(maybe_int)
        return True

    except:
        return False


def avg(items: list):
    try:
        return round(sum(items) / len(items), 5)

    except AttributeError as t:
        print("Requires list type.")
        with open("error_log.txt", "a+") as fe:
            err = f"Requires list type {str(datetime.datetime.now())}\n"
            fe.write(err)

    except TypeError as e:
        print("Could not compute the average.")
        with open("error_log.txt", "a+") as fe:
            err = f"Could not compute the average {str(datetime.datetime.now())}\n"
            fe.write(err)


def most_frequent(items: list):
    try:
        return Counter(items).most_common(1)[0][0]

    except AttributeError as t:
        print("Requires list type.")
        with open("error_log.txt", "a+") as fe:
            err = f"Requires list type {str(datetime.datetime.now())}\n"
            fe.write(err)

    except BaseException as e:
        print("Cannot find most frequent item.")
        with open("error_log.txt", "a+") as fe:
            err = f"Cannot find most frequent item {str(datetime.datetime.now())}\n"
            fe.write(err)


def get_temp(items: list) -> tuple:
    try:
        items.sort()
        min_temp = items[0]
        max_temp = items[-1]
        avg_temp = avg(items)

        return min_temp, max_temp, avg_temp

    except AttributeError as t:
        print("Requires list type.")
        with open("error_log.txt", "a+") as fe:
            err = f"Requires list type {str(datetime.datetime.now())}\n"
            fe.write(err)

    except BaseException as e:
        print("Could not find temperature values.")
        with open("error_log.txt", "a+") as fe:
            err = f"Could not find temperature values {str(datetime.datetime.now())}\n"
            fe.write(err)


def weather_info(year: int, month: int):
    url1 = f'https://www.timeanddate.com/weather/iran/tehran/historic?month={month}&year={year}'
    url2 = f"https://www.accuweather.com/en/ir/tehran/210841/{months[month]}-weather/210841?year={year}&view=list"

    driver.get(url1)

    select_div = Select(driver.find_element_by_id('wt-his-select'))
    options_len = len(select_div.options)
    days = [int(s.text.split(" ")[0]) for s in select_div.options]

    daily_data = {}
    for day in range(options_len):
        time.sleep(3)

        temperature, weather, wind = day_info((year, month, days[day]))

        data = {
            "day": days[day],
            "lowest_temp": temperature[0],
            "highest_temp": temperature[1],
            "avg_temp": temperature[2],
            "weather_status": weather,
            "wind_power": wind,
        }
        daily_data[days[day]] = data

        if day != options_len - 1:
            select_div.select_by_index(day + 1)

        dates.append(str(datetime.date(year, month, days[day])))
        lowest_temp_data.append(temperature[0])
        highest_temp_data.append(temperature[1])
        avg_temp_data.append(temperature[2])
        weather_status_data.append(weather)
        wind_power_data.append(wind)

        print(f"{by}{year}/{month}/{days[day]} done{end}")

    result[year][month] = daily_data

    # driver.find_element_by_tag_name('body').send_keys(Keys.CONTROL + 't')

    driver.get(url2)
    time.sleep(5)

    current_date = datetime.datetime.utcnow()
    if year == current_date.year and month == current_date.month:
        precip = get_precip(days_limit=current_date.day)
    else:
        precip = get_precip()

    for _day in range(options_len):
        result[year][month][days[_day]]["precipitation"] = precip[days[_day] - 1]
        precipitation_data.append(precip[days[_day] - 1])


def get_precip(days_limit=100) -> list:
    res = []

    def get_info():
        # rains = driver.find_elements_by_class_name("show-tablet")
        rains = WebDriverWait(driver, delay).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "show-tablet")))

        i = 1
        for rain in rains:
            if i == days_limit:
                break

            out = rain.get_attribute('innerText')
            out = out.replace('\n', '')
            out = out.replace('\t', '')
            out = out.replace(' ', '')
            out = float(out[6:-2])

            res.append(out)

            i += 1

    try:
        time.sleep(5)
        get_info()

        if res is None or not (len(res) > 0):
            print(f"{rb}Wait a bit longer....{end}")

            time.sleep(5)
            get_info()

        return res

    except BaseException as e:
        print("Error in precipitation.")
        print(e)
        with open("error_log.txt", "a+") as fe:
            err = f"Error in precipitation {str(datetime.datetime.now())}\n"
            fe.write(err)
        
        print(f"{rb}Wait a bit longer.....{end}")

        driver.refresh()
        time.sleep(5)
        get_info()


def day_info(date_info: tuple) -> tuple:
    temp = []
    weather = []
    wind = []

    # table = driver.find_element_by_id("wt-his")
    def get_info():
        table = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.ID, "wt-his")))

        for t in table.find_elements(By.TAG_NAME, "tr"):
            k = 0
            for d in t.find_elements(By.TAG_NAME, "td"):
                if k == 1:
                    tmp = d.text[:-3]
                    if tmp == "" or not is_int(tmp):
                        break
                    temp.append(int(tmp))
                if k == 2:
                    weather.append(d.text[:-1])
                if k == 3:
                    item = d.text[:-5]
                    if (item == "" or not is_int(item)) and item != "No":
                        break
                    elif item == "No":
                        wind.append(0)
                    else:
                        wind.append(int(item))
                k += 1

    try:
        get_info()

    except (NoSuchElementException, StaleElementReferenceException):
        print(f"{rb}Wait a bit longer...{end}")

        with open("error_log.txt", "a+") as fe:
            err = f"information of {date_info[0]}/{date_info[1]}/{date_info[2]} did not load completely in first try at {str(datetime.datetime.now())}\n"
            fe.write(err)

        time.sleep(5)

        get_info()

    return get_temp(temp), most_frequent(weather), avg(wind)


def get_all_data(start_year: int, end_year: int):
    current_date = datetime.datetime.utcnow()
    months_of_year = 12

    if end_year > current_date.year:
        end_year = current_date.year

    if start_year < 2010:
        start_year = 2010

    with open("error_log.txt", "a+") as fe:
        err = f"===============================\nstarted at {str(datetime.datetime.now())}\n" \
              f"-------------------------------\n"
        fe.write(err)

    for y in range(start_year, end_year + 1):
        result[y] = {}
        if y == current_date.year:
            months_of_year = current_date.month
        for m in range(1, months_of_year + 1):
            if y == end_year and m > current_date.month:
                break

            weather_info(y, m)
            print(f"{yb}{y}/{m} completed{end}")

        with open("weather_log.json", "w+") as fw:
            fw.write(json.dumps(result))

    with open("error_log.txt", "a+") as fe:
        err = f"-------------------------------\nfinished at {str(datetime.datetime.now())}\n" \
              f"===============================\n"
        fe.write(err)

    driver.close()
    driver.quit()

    excel_writer()

    return json.dumps(result)

def get_yesterday():
    current_date = datetime.datetime.utcnow()

    url1 = "https://www.timeanddate.com/weather/iran/tehran/historic"
    url2 = f"https://www.accuweather.com/en/ir/tehran/210841/{months[current_date.month]}-weather/210841?year={current_date.year}&view=list"

    driver.get(url1)

    select_div = Select(driver.find_element_by_id('wt-his-select'))
    select_div.select_by_index(2)

    time.sleep(3)
    temperature, weather, wind = day_info((current_date.year, current_date.month, current_date.day - 1))
    
    driver.get(url2)

    precip = get_precip(days_limit=current_date.day)

    dates.append(str(datetime.date(current_date.year, current_date.month, current_date.day - 1)))
    lowest_temp_data.append(temperature[0])
    highest_temp_data.append(temperature[1])
    avg_temp_data.append(temperature[2])
    weather_status_data.append(weather)
    wind_power_data.append(wind)
    precipitation_data.append(precip[-1])

    driver.close()
    driver.quit()

    data = {
            "dates": dates,
            "lowest_temp_data": lowest_temp_data,
            "highest_temp_data": highest_temp_data,
            "avg_temp_data": avg_temp_data,
            "weather_status_data": weather_status_data,
            "wind_power_data": wind_power_data,
            "precipitation_data": precipitation_data,
        }
	
    print(gb)
    print(data)
    print(end)
    # excel_writer()


if __name__ == "__main__":
    crawl_type = int(sys.argv[1])

    if crawl_type == 0:
        start_input = int(sys.argv[2])
        end_input = int(sys.argv[3])

        get_all_data(start_input, end_input)

        # schedule.every().day.at("01:00").do(get_all_data(start_input, end_input))
        #
        # while True:
        #     schedule.run_pending()
        #     time.sleep(5)

    elif crawl_type == 1:
        # get_yesterday()

        # schedule.every().day.at("01:00").do(get_yesterday())
        schedule.every(10).minutes.do(get_yesterday)
        
        while True:
            schedule.run_pending()
            time.sleep(5)
